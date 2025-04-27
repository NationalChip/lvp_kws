#!/usr/bin/env python3
# -*- coding: utf-8 -*-



import sys, re
from openpyxl import load_workbook
try:
    from modules.common.debugLogger import DebugLogger
except:
    sys.path.append("../../")
    from modules.common.debugLogger import DebugLogger

__version__="1.0.0"

class GetWordThreshold:
    """ 获取词阈值处理类"""
    def __init__(self, logger=None, internal_msg_bus={}):
        self.skip_list = []
        self.logger = logger
        if not self.logger:
            self.logger = DebugLogger()
        self.internal_msg_bus = internal_msg_bus

    def get_keyword_info(self, keywords_filename):
        """
            功能:
                    * 读取 BunKws 模型输出包中的 keyword.txt
                    * 根据该文件中指定的顺序组成有一个有序的 关键字字典
            参数:
                    * str: keyword.txt 文件路径 keywords_filename
            返回值:
                    * dict: 有序关键字列表
        """
        fail_flag = 0
        output_dict = {}
        with open(keywords_filename, "r") as f:
            data = f.readlines()
        for single_line in data:
            single_line = single_line.replace("  ", " ").replace("\n", "")

            tmp_dict = {}
            if re.findall("是否主唤醒:", single_line):
                is_main_wakeup = single_line.split("是否主唤醒:")[-1].strip()
                if str(is_main_wakeup) not in ["0", "1"]:
                    self.logger.log(f"[GetWordThreshold]: 错误！{keywords_filename} 文件中 {single_line} 是否为主唤醒设置错误，该值只能是0/1", "error")
                    fail_flag = 1
                tmp_dict["是否主唤醒"] = is_main_wakeup
            if re.findall("事件id:", single_line):
                event_id = 100
                if "是否主唤醒" in tmp_dict:
                    event_id = single_line.split("是否主唤醒:")[0].split("事件id:")[-1].strip()
                else:
                    event_id = single_line.split("事件id:")[-1].strip()
                try:
                    event_id = int(event_id)
                except:
                    self.logger.log(f"[GetWordThreshold]: 错误！{keywords_filename} 文件中 {single_line} 事件ID 设置错误，该值只能是整数", "error")
                    fail_flag = 1
                tmp_dict["事件ID"] = event_id
            index = int(single_line.split("是否主唤醒:")[0].split("事件id:")[0].strip().split(" ")[-1].strip())
            #index = int(single_line.split(" ")[-1].strip())
            #print(index)
            if index in output_dict:
                self.logger.log(f"[GetWordThreshold]: 错误！{keywords_filename} 文件中有序索引 {index} 有重复", "error")
                fail_flag = 1
            if index == -1:
                continue
            cmd = single_line.split(" ")[0].strip()
            if cmd == "<filler>" and index == "-1":
                continue
            output_dict[index] = {cmd: tmp_dict}
        if fail_flag:
            return {}
        output_dict = {k: output_dict[k] for k in sorted(output_dict)}
        self.logger.log(f"[GetWordThreshold]: 调试！{keywords_filename} 文件解析后的数据 {output_dict}", "debug")
        return output_dict


    def GetThresholdData(self):
        """
            功能:
                    *
        """
        keywords_filename = self.internal_msg_bus["输入文件"]["词有序列表文件"]
        report_filename = self.internal_msg_bus["输入文件"]["模型报告文件"]
        keywords_dict = self.get_keyword_info(keywords_filename)
        #print(keywords_dict)
        if not keywords_dict:
            self.logger.log(f"[GetWordThreshold]: 错误！未正确从 {keywords_filename} 文件中获取到有序关键字列表信息", "error")
            return {}

        # 加载 Excel 文件
        wb = load_workbook(filename=report_filename)
        # 选择工作表（默认第一个或指定名称）
        sheet = wb.active  # 或 sheet = wb['Sheet1']
        #找到指令词所在的列
        cmd_col = 0 #默认第0列
        try:
            cmd_col = column_names.index("threshold")
        except:
            pass
        #找到阈值所在的列
        threshold_col = 1 #默认第1列
        try:
            threshold_col = column_names.index("threshold")
        except:
            pass
        cmd_threshold_dict = {}
        cmd_min_row = 4
        threshold_row = cmd_min_row
        for row in sheet.iter_rows(min_row=cmd_min_row, max_col=cmd_col, min_col=cmd_col, values_only=True):
            cmd = row[0]
            threshold = sheet.cell(row=threshold_row, column=threshold_col+1).value
            try:
                if threshold < 1:
                    threshold = round(float(threshold*1000))
                else:
                    threshold = round(float(threshold*10))
            except Exception as e:
                self.logger.log(f"[GetWordThreshold]: 错误！未正确从 {report_filename} 文件中获取到 {cmd} 的阈值信息", "error")
            threshold_row += 1
            if cmd in ["None", None]:
                continue
            cmd_threshold_dict[cmd] = threshold
        #排序及检查
        output_dict = {}
        fail_flag = 0
        for keywords_index in keywords_dict:
            cmd = list(keywords_dict[keywords_index].keys())[0]
            if cmd not in cmd_threshold_dict:
                self.logger.log(f"[GetWordThreshold]: 错误！未正确从 {cmd} 关键字阈值信息未正确从 {report_filename} 中获取", "error")
                fail_flag = 1

            event_id = 100
            if "事件ID" in keywords_dict[keywords_index][cmd]:
                event_id = keywords_dict[keywords_index][cmd]["事件ID"]
            else:
                self.logger.log(f"[GetWordThreshold]: 提示！ 词: {cmd}  未在 {keywords_filename} 中配置 '事件ID'，将默认为 {event_id}", "info")

            is_main_wakeup = 1
            if "是否主唤醒" in keywords_dict[keywords_index][cmd]:
                is_main_wakeup = keywords_dict[keywords_index][cmd]["是否主唤醒"]
            else:
                self.logger.log(f"[GetWordThreshold]: 提示！ 词: {cmd}  未在 {keywords_filename} 中配置 '是否主唤醒'，将默认为主唤醒", "info")

            if not fail_flag:
                output_dict[keywords_index] = {"词": cmd, "阈值": cmd_threshold_dict[cmd], "事件ID": event_id, "是否主唤醒": is_main_wakeup}
        #print(output_dict)
        # 关闭工作簿（非必需，但建议）
        wb.close()
        if fail_flag:
            return {}
        #self.logger.log(output_dict, "debug")
        return output_dict
